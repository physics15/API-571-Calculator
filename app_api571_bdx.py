import io
import json
import re
import shutil
import subprocess
import tempfile
from pathlib import Path

import openpyxl
import pandas as pd
import streamlit as st
from PyPDF2 import PdfReader

st.set_page_config(page_title='API 571 BDX Predictor', layout='wide')

LOOP_DEFAULTS = {
    1: dict(temp=35, min_temp=20, max_temp=60, pressure=6, aqueous='No', hydrocarbon='Yes', gas='No', condense='No', wetting='Continuous', insulation='Yes', insulation_condition='Unknown', tan=0.1, deadlegs='Yes', deposits='Yes', notes='1,3-butadiene loop; organic acid corrosion, CUI, under-deposit concerns, popcorn polymer risk in document.'),
    2: dict(temp=35, min_temp=20, max_temp=60, pressure=8, aqueous='No', hydrocarbon='Yes', gas='No', condense='No', wetting='Continuous', insulation='Yes', insulation_condition='Unknown', tan=0.1, deadlegs='No', deposits='No', notes='C4 hydrocarbon loop; mostly non-corrosive hydrocarbons, CUI remains possible.'),
    3: dict(temp=50, min_temp=30, max_temp=80, pressure=4, aqueous='Yes', hydrocarbon='Yes', gas='No', condense='Yes', wetting='Continuous', insulation='Yes', insulation_condition='Unknown', tan=0.1, deadlegs='Yes', deposits='Yes', notes='Lean solvent loop; aqueous + hydrocarbon mixed service with potential organic acid corrosion.'),
    4: dict(temp=55, min_temp=35, max_temp=90, pressure=4, aqueous='Yes', hydrocarbon='Yes', gas='No', condense='Yes', wetting='Continuous', insulation='Yes', insulation_condition='Unknown', tan=0.1, deadlegs='Yes', deposits='Yes', notes='Rich solvent loop; organic acid corrosion, nitrite/nitrate-related concerns, fouling/under-deposit concerns.'),
    5: dict(temp=45, min_temp=25, max_temp=70, pressure=3, aqueous='Yes', hydrocarbon='Yes', gas='No', condense='Yes', wetting='Continuous', insulation='Yes', insulation_condition='Unknown', tan=0.1, deadlegs='No', deposits='Yes', notes='Water-washed hydrocarbon/service area with organic acid and under-deposit concerns.'),
    6: dict(temp=30, min_temp=10, max_temp=50, pressure=2, aqueous='No', hydrocarbon='Yes', gas='No', condense='No', wetting='Continuous', insulation='No', insulation_condition='N/A', tan=None, deadlegs='No', deposits='No', notes='Methanol service; document notes methanol SCC under very dry conditions plus organic acid corrosion.'),
    7: dict(temp=35, min_temp=20, max_temp=50, pressure=2, aqueous='Yes', hydrocarbon='No', gas='No', condense='No', wetting='Continuous', insulation='No', insulation_condition='N/A', tan=None, deadlegs='No', deposits='No', notes='Sodium nitrite solution; nitric acid corrosion is document-led but not directly represented in the workbook.'),
    8: dict(temp=40, min_temp=25, max_temp=60, pressure=2, aqueous='Yes', hydrocarbon='No', gas='No', condense='Yes', wetting='Continuous', insulation='No', insulation_condition='N/A', tan=0.1, deadlegs='No', deposits='No', notes='Process wash water; pH window is 7 to 10; solvent decomposition can form acetic acid.'),
    9: dict(temp=-20, min_temp=-40, max_temp=10, pressure=16.7, aqueous='No', hydrocarbon='Yes', gas='No', condense='No', wetting='Continuous', insulation='Yes', insulation_condition='Unknown', tan=None, deadlegs='No', deposits='No', notes='Propylene/LPG cooling medium; brittle fracture screening should be retained because of low temperature.'),
    10: dict(temp=30, min_temp=20, max_temp=50, pressure=1, aqueous='No', hydrocarbon='No', gas='Yes', condense='No', wetting='Intermittent', insulation='No', insulation_condition='N/A', tan=None, deadlegs='No', deposits='No', notes='Vent gas line exposed to atmosphere; atmospheric corrosion is the main document-led mechanism.'),
    11: dict(temp=40, min_temp=20, max_temp=60, pressure=2, aqueous='No', hydrocarbon='Yes', gas='Yes', condense='Yes', wetting='Intermittent', insulation='No', insulation_condition='N/A', tan=0.1, deadlegs='No', deposits='No', notes='Off-gas / flare-connected line; hydrocarbons with possible water vapour fractions.'),
    12: dict(temp=30, min_temp=15, max_temp=45, pressure=1, aqueous='Yes', hydrocarbon='Yes', gas='No', condense='Yes', wetting='Continuous', insulation='No', insulation_condition='N/A', tan=0.1, deadlegs='No', deposits='No', notes='Drips, drains, slops; mixed waste streams with reported pit-like attack.'),
}


def normalize_space(s: str) -> str:
    return re.sub(r'\s+', ' ', s).strip()


def extract_loop_sections(pdf_path: Path):
    reader = PdfReader(str(pdf_path))
    page_texts = {i + 1: reader.pages[i].extract_text() or '' for i in range(len(reader.pages))}
    loop_pages = ''.join(page_texts[p] + '\n' for p in range(17, 35) if p in page_texts)
    pattern = re.compile(r'(4\.(\d+)\.\s+Corrosion Loop #\s*(\d+)\s*[–-]\s*(.+?))(?=4\.\d+\.\s+Corrosion Loop #|References\s|\Z)', re.S)
    sections = []
    for full, sec_no, loop_no, title_and_body in pattern.findall(loop_pages):
        loop_no = int(loop_no)
        title_and_body = title_and_body.strip()
        title = normalize_space(title_and_body.split('\n', 1)[0])
        body = title_and_body[len(title_and_body.split('\n', 1)[0]):].strip()
        if 1 <= loop_no <= 12:
            sections.append({
                'loop_no': loop_no,
                'section_no': sec_no,
                'title': title,
                'body': body,
                'full_text': normalize_space(title + ' ' + body),
            })
    dedup = {}
    for sec in sections:
        dedup[sec['loop_no']] = sec
    return [dedup[i] for i in sorted(dedup)]


def extract_mechanism_bullets(section_text: str):
    m = re.search(r'Potential\s+Corrosion/Degradation/Fouling\s+Mechanisms\s*(.*?)(Integrity\s+Operating\s+Windows|Special\s+Inspection\s+Considerations|Special\s+Start)', section_text, re.I | re.S)
    if not m:
        return []
    block = m.group(1)
    parts = re.split(r'\s*[•]\s*', block)
    return [normalize_space(p) for p in parts if normalize_space(p)]


def infer_inputs(loop_no: int, title: str, full_text: str):
    d = LOOP_DEFAULTS[loop_no]
    bullets = extract_mechanism_bullets(full_text)
    notes = bullets[:] + ([d['notes']] if d['notes'] else [])
    values = {
        'B4': 'Initial Screening (No Inspection)',
        'B7': 'Other', 'D7': 'Butadiene Extraction Unit',
        'B8': 'Piping', 'D8': 'Loop-level screening from CCM text',
        'B9': f'Corrosion Loop #{loop_no}', 'B10': title,
        'B14': 'Carbon Steel', 'B15': 'Generic loop default from CCM unless stated otherwise',
        'B16': 'None', 'B17': 'Unknown', 'B18': 'Yes',
        'B20': d['insulation'], 'B21': d['insulation_condition'],
        'B40': 'Medium', 'B41': 'No', 'B42': 'No', 'B47': 'Unknown',
        'B51': 'No', 'B52': 'No', 'B55': d['temp'], 'B56': d['min_temp'], 'B57': d['max_temp'],
        'B58': d['temp'], 'B59': d['pressure'],
        'B60': 'Yes' if ('thermal fatigue' in full_text.lower() or 'temperature difference' in full_text.lower()) else 'No',
        'B61': d['aqueous'], 'B62': d['hydrocarbon'], 'B63': d['gas'], 'B64': d['condense'],
        'B67': d['wetting'], 'B68': 'No', 'B78': 'No', 'B79': 'No', 'B80': 'No', 'B81': 'No',
        'B82': 'Yes' if 'oxygen' in full_text.lower() or 'peroxide' in full_text.lower() else 'No',
        'B85': 'Yes' if loop_no == 6 else 'No', 'B88': 'No', 'B89': 'N/A',
        'B97': 'No', 'B98': 'No', 'B101': 'Yes' if loop_no == 6 else 'No',
        'B103': 'No', 'B105': 'No', 'B111': 1.0,
        'B112': 'Two phase' if d['aqueous'] == 'Yes' and d['hydrocarbon'] == 'Yes' else ('Mist' if d['gas'] == 'Yes' and d['condense'] == 'Yes' else 'Single phase'),
        'B113': 'No', 'B114': d['deadlegs'], 'B115': d['deposits'],
        'B119': 'No', 'B121': 'None', 'B122': 'None', 'B125': 'No', 'B126': 'Unknown',
        'B129': 'Auto-generated from BDX CCM text; no direct inspection evidence loaded.',
        'B130': 'No', 'B131': 'No', 'B132': 'No', 'B133': 'No', 'B134': 'No', 'B135': 'No', 'B136': 'No', 'B137': 'No', 'B138': 'No', 'B139': 'No',
        'B108': ' | '.join(notes)[:30000],
    }
    if d['tan'] is not None:
        values['B77'] = d['tan']
    if loop_no in {7, 8}:
        values['B73'] = 8
    if loop_no == 8:
        values['B15'] = 'C-steel default; duplex SS2205 also noted for Ex822/Ex823'
    if loop_no == 3:
        values['B61'] = 'Yes'; values['B73'] = 7; values['B114'] = 'No'
    if loop_no == 4:
        values['B61'] = 'Yes'; values['B73'] = 7
    if loop_no == 5:
        values['B73'] = 7
    if loop_no == 6:
        values['B82'] = 'Yes'; values['B61'] = 'No'
    if loop_no == 7:
        values['B73'] = 8; values['B82'] = 'Yes'
    if loop_no == 9:
        values['B63'] = 'No'; values['B65'] = -40
    if loop_no == 10:
        values['B63'] = 'Yes'; values['B61'] = 'No'; values['B62'] = 'No'
    if loop_no == 11:
        values['B63'] = 'Yes'; values['B64'] = 'Yes'
    if loop_no == 12:
        values['B61'] = 'Yes'; values['B62'] = 'Yes'

    lowered = full_text.lower()
    if 'cui' in lowered:
        values['B20'] = 'Yes'; values['B21'] = 'Unknown'
    if 'underdeposit corrosion' in lowered or 'under deposit corrosion' in lowered or 'deposits' in lowered:
        values['B115'] = 'Yes'
    if 'deadleg' in lowered:
        values['B114'] = 'Yes'
    if 'water level' in lowered or 'condensate' in lowered:
        values['B64'] = 'Yes'
    if 'organic acid corrosion' in lowered and 'B77' not in values:
        values['B77'] = 0.1
    if 'nitric acid corrosion' in lowered:
        values['B108'] = (values.get('B108', '') + ' | Document-specific gap: nitric acid corrosion is not directly named in stock API 571 calculator database.').strip(' |')[:30000]
    if 'methanol scc' in lowered:
        values['B108'] = (values.get('B108', '') + ' | Document-specific gap: methanol SCC is approximated through ethanol SCC surrogate flags.').strip(' |')[:30000]
    return values


def populate_workbook(template_path: Path, out_path: Path, cell_values: dict):
    wb = openpyxl.load_workbook(template_path)
    ws = wb['INPUT']
    for cell, value in cell_values.items():
        ws[cell] = value
    wb.save(out_path)


def recalc_with_libreoffice(xlsx_path: Path):
    calc_dir = xlsx_path.parent / '_calc_cache'
    calc_dir.mkdir(parents=True, exist_ok=True)
    try:
        subprocess.run([
            'libreoffice', '--headless', '--convert-to', 'xlsx', '--outdir', str(calc_dir), str(xlsx_path)
        ], check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        recalced = calc_dir / xlsx_path.name
        if recalced.exists():
            shutil.move(str(recalced), str(xlsx_path))
        return True, None
    except Exception as e:
        return False, str(e)


def extract_results(xlsx_path: Path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['RESULTS']
    top = []
    for r in range(7, 27):
        mech = ws[f'C{r}'].value
        score = ws[f'D{r}'].value
        if mech and score not in (None, '', 0):
            top.append({
                'rank': ws[f'A{r}'].value,
                'api571_section': ws[f'B{r}'].value,
                'mechanism': mech,
                'likelihood_pct': round(float(score) * 100, 2) if isinstance(score, (int, float)) and score <= 1 else score,
                'confidence': ws[f'E{r}'].value,
            })
    warnings = []
    for r in range(29, 48):
        val = ws[f'A{r}'].value
        if isinstance(val, str) and val.strip():
            warnings.append(val.strip())
    return {
        'summary': ws['A3'].value,
        'mode': ws['A4'].value,
        'top': top,
        'warnings': warnings,
    }


def run_pipeline(pdf_bytes: bytes, wb_bytes: bytes):
    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        pdf_path = td / 'ccm.pdf'
        xlsx_path = td / 'API571_Calculator.xlsx'
        pdf_path.write_bytes(pdf_bytes)
        xlsx_path.write_bytes(wb_bytes)

        sections = extract_loop_sections(pdf_path)
        all_rows = []
        per_loop = {}
        workbooks = {}
        for sec in sections:
            values = infer_inputs(sec['loop_no'], sec['title'], sec['full_text'])
            loop_xlsx = td / f"Loop_{sec['loop_no']:02d}.xlsx"
            populate_workbook(xlsx_path, loop_xlsx, values)
            recalced, err = recalc_with_libreoffice(loop_xlsx)
            results = extract_results(loop_xlsx)
            results['recalculated'] = recalced
            results['recalc_error'] = err
            results['inputs'] = values
            results['title'] = sec['title']
            results['loop_no'] = sec['loop_no']
            per_loop[sec['loop_no']] = results
            workbooks[f'Loop_{sec["loop_no"]:02d}.xlsx'] = loop_xlsx.read_bytes()
            for row in results['top'][:5]:
                all_rows.append({
                    'loop_no': sec['loop_no'],
                    'loop_title': sec['title'],
                    **row,
                })
        df = pd.DataFrame(all_rows)
        summary_json = json.dumps(per_loop, indent=2, default=str)
        return sections, per_loop, df, summary_json, workbooks


st.title('API 571 Butadiene Damage Mechanism Predictor')
st.caption('Upload the CCM PDF and the API 571 calculator workbook. The app reads the corrosion loops, auto-fills the INPUT sheet, recalculates, and shows ranked mechanisms.')

with st.sidebar:
    st.header('Inputs')
    pdf_file = st.file_uploader('CCM PDF', type=['pdf'])
    wb_file = st.file_uploader('API 571 Calculator Workbook', type=['xlsx'])
    use_defaults = st.checkbox('Use bundled defaults if available', value=True)
    run_btn = st.button('Run prediction', type='primary')


default_pdf = Path('/mnt/data/CCM Butadiene Extraction Units_Mar20.pdf')
default_wb = Path('/mnt/data/API571_Calculator_v7.xlsx')

if run_btn:
    try:
        pdf_bytes = pdf_file.read() if pdf_file else (default_pdf.read_bytes() if use_defaults and default_pdf.exists() else None)
        wb_bytes = wb_file.read() if wb_file else (default_wb.read_bytes() if use_defaults and default_wb.exists() else None)
        if not pdf_bytes or not wb_bytes:
            st.error('Please upload both the CCM PDF and the calculator workbook, or enable bundled defaults.')
        else:
            with st.spinner('Running loop extraction, autofill, and scoring...'):
                sections, per_loop, df, summary_json, workbooks = run_pipeline(pdf_bytes, wb_bytes)

            c1, c2, c3 = st.columns(3)
            c1.metric('Loops found', len(sections))
            c2.metric('Top-5 rows', len(df))
            c3.metric('Workbook outputs', len(workbooks))

            st.subheader('Top-ranked mechanisms by loop')
            st.dataframe(df, use_container_width=True)

            csv_bytes = df.to_csv(index=False).encode('utf-8')
            st.download_button('Download Top-5 CSV', data=csv_bytes, file_name='BDX_API571_Top5.csv', mime='text/csv')
            st.download_button('Download Summary JSON', data=summary_json.encode('utf-8'), file_name='BDX_API571_AutoInput_Summary.json', mime='application/json')

            st.subheader('Loop details')
            loop_options = [f"Loop {i:02d} - {per_loop[i]['title']}" for i in sorted(per_loop)]
            selected = st.selectbox('Select loop', loop_options)
            selected_no = int(selected.split()[1])
            data = per_loop[selected_no]

            left, right = st.columns([2, 1])
            with left:
                st.markdown(f"### Loop {selected_no:02d}: {data['title']}")
                st.write(data.get('summary'))
                st.dataframe(pd.DataFrame(data['top']), use_container_width=True)
                if data['warnings']:
                    st.markdown('#### Workbook warnings')
                    for w in data['warnings']:
                        st.write(f'- {w}')
            with right:
                st.markdown('#### Recalculation status')
                st.write('Success' if data['recalculated'] else 'Fallback / not recalculated')
                if data['recalc_error']:
                    st.code(data['recalc_error'])
                st.markdown('#### Download workbook')
                fname = f'Loop_{selected_no:02d}.xlsx'
                st.download_button('Download populated workbook', data=workbooks[fname], file_name=fname, mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

            with st.expander('Show auto-filled INPUT values for selected loop'):
                inp_df = pd.DataFrame(sorted(data['inputs'].items()), columns=['Cell', 'Value'])
                st.dataframe(inp_df, use_container_width=True)

            with st.expander('Show extracted loop sections'):
                sec_df = pd.DataFrame([{'loop_no': s['loop_no'], 'title': s['title'], 'section_no': s['section_no']} for s in sections])
                st.dataframe(sec_df, use_container_width=True)

    except Exception as e:
        st.exception(e)
else:
    st.info('Choose the files in the sidebar, then click Run prediction.')
    if use_defaults and default_pdf.exists() and default_wb.exists():
        st.success('Bundled defaults detected. You can run the app immediately without uploading files.')
