import os
import re
import json
import pandas as pd
from PyPDF2 import PdfReader
from openpyxl import load_workbook

# =========================
# 1. EXTRACT TEXT FROM PDF
# =========================
def extract_text(pdf_path):
    reader = PdfReader(pdf_path)
    text = ""
    for page in reader.pages:
        text += page.extract_text() + "\n"
    return text


# =========================
# 2. IDENTIFY CORROSION LOOPS
# =========================
def extract_loops(text):
    pattern = r"(Corrosion Loop\s*\d+.*?)(?=Corrosion Loop\s*\d+|$)"
    loops = re.findall(pattern, text, re.DOTALL)
    return loops


# =========================
# 3. SIMPLE RULE-BASED PARSER
# =========================
def infer_inputs(loop_text):
    inputs = {
        "Material": "Carbon Steel",
        "Temperature": "Ambient",
        "Chemistry": "",
        "Flow": "Normal",
        "Evidence": "None"
    }

    if "acid" in loop_text.lower():
        inputs["Chemistry"] = "Acidic"
    if "methanol" in loop_text.lower():
        inputs["Chemistry"] = "Methanol"
    if "water" in loop_text.lower():
        inputs["Chemistry"] = "Wet"
    if "high temperature" in loop_text.lower():
        inputs["Temperature"] = "High"
    if "deadleg" in loop_text.lower():
        inputs["Flow"] = "Stagnant"
    if "corrosion" in loop_text.lower():
        inputs["Evidence"] = "General Corrosion"

    return inputs


# =========================
# 4. POPULATE EXCEL
# =========================
def fill_excel(template_path, inputs, output_path):
    wb = load_workbook(template_path)
    ws = wb["INPUT"]

    ws["B2"] = inputs["Material"]
    ws["B3"] = inputs["Temperature"]
    ws["B4"] = inputs["Chemistry"]
    ws["B5"] = inputs["Flow"]
    ws["B6"] = inputs["Evidence"]

    wb.save(output_path)


# =========================
# 5. MAIN PIPELINE
# =========================
def run_pipeline(pdf_path, excel_template, output_dir="outputs"):
    os.makedirs(output_dir, exist_ok=True)

    text = extract_text(pdf_path)
    loops = extract_loops(text)

    results = []

    for i, loop in enumerate(loops, 1):
        inputs = infer_inputs(loop)

        output_excel = os.path.join(output_dir, f"loop_{i}.xlsx")
        fill_excel(excel_template, inputs, output_excel)

        results.append({
            "loop": i,
            "inputs": inputs,
            "file": output_excel
        })

    # save summary
    with open(os.path.join(output_dir, "summary.json"), "w") as f:
        json.dump(results, f, indent=2)

    return results
